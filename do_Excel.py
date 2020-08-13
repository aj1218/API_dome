from openpyxl import load_workbook
from read_config import ReadConfig
import project_path
from get_data import GetCookie
from do_regx import DoRegs
import os
import pdb
class Doexcel:
    @classmethod
    def get_excel(cls,file_name):
        wb = load_workbook(file_name)  # linux路径
        test_data=[]  #拿到字典里面的所有数据

        mode=eval(ReadConfig.get_config(project_path.case_config_path,'MODE','mode'))
        # tel=getattr(GetCookie,'NoRegTel')#利用反射拿到数据
        normal_tel=getattr(GetCookie,"normal_tel")  #从getdata里面直接拿数据 利用反射
        for key in mode:  #遍历配置文件里面的字典
            sheet = wb[key]  #key就是表单名
            if mode[key] == 'all':
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data["case_id"]=sheet.cell(i,1).value  #行号
                    row_data["url"]=sheet.cell(i,2).value
                    # row_data["data"] = sheet.cell(i, 3).value    #拿到手机号之后这个地方要做替换
                    # if sheet.cell(i,3).value.find('${normal_tel}')!= -1: #找这一行的第三列看S{tel_1}的值
                    #     # row_data['data']=sheet.cell(i, 3).value.replace('${normal_tel}',str(normal_tel))  #字符串替换
                    #     row_data['data'] = DoRegs.do_regs(sheet.cell(i, 3).value)
                    if sheet.cell(i,3).value.find('${normal_tel}')!=-1:
                        row_data['data'] = sheet.cell(i,3).value.replace('${normal_tel}',str(normal_tel+1))
                    else:  #如果没有拿到值
                        # row_data['data'] = sheet.cell(i, 3).value
                        row_data['data'] = DoRegs.do_regs(sheet.cell(i, 3).value)

                    # if sheet.cell(i, 4).value ==None:
                    #     Doexcel().write_back(project_path.test_case_path, "invest", row_data["case_id"]+1,4,str(normal_tel))
                    if sheet.cell(i, 4).value != None:
                        if sheet.cell(i, 4).value.find('${normal_tel}')!=-1:
                            # row_data['check_sql'] = DoRegs.do_regs(sheet.cell(i, 4).value)
                            # row_data["check_sql"]=sheet.cell(i, 4).value.replace('${normal_tel}',str(normal_tel))
                            row_data['check_sql'] = DoRegs.do_regs(sheet.cell(i, 4).value)

                    else:
                        row_data["check_sql"] = sheet.cell(i, 4).value
                    row_data["expected"] = sheet.cell(i, 5).value #添加一个期望值到测试数据里面去
                    row_data["method"] = sheet.cell(i, 6).value
                    row_data["title"] = sheet.cell(i, 7).value
                    row_data["sheet_name"]=key
                    test_data.append(row_data)
                    cls.update_tel(normal_tel+2,file_name,"init") #更新手机号
            else:
                for case_id in mode[key]: #mode可以不等于 all的时候就可以等于他的列表里面的数据
                    row_data = {} #字典
                    print(case_id)
                    row_data["case_id"] = sheet.cell(case_id+1, 1).value # 行号
                    row_data["url"] = sheet.cell(case_id+1, 2).value
                    if sheet.cell(case_id+1,3).value.find('${normal_tel}')!=-1:
                        row_data['data'] = sheet.cell(case_id+1,3).value.replace('${normal_tel}',str(normal_tel+1))
                    else:  #如果没有拿到值
                        # row_data['data'] = sheet.cell(i, 3).value
                        row_data['data'] = DoRegs.do_regs(sheet.cell(case_id+1, 3).value)
                    # if sheet.cell(case_id+1,3).value.find('${normal_tel}')!= -1: #找这一行的第三列看S{tel_1}的值
                    #     row_data['data']=sheet.cell(case_id+1, 3).value.replace('${normal_tel}',str(normal_tel))  #字符串替换
                    #
                    # elif sheet.cell(case_id+1,3).value.find('${normal_tel1}')!=-1:
                    #     row_data['data'] = sheet.cell(case_id+1,3).value.replace('${normal_tel1}',str(normal_tel))
                    # else:  #如果没有拿到值
                    #     row_data["data"] = sheet.cell(case_id+1, 3).value
                      #这是sql语句的处理
                    if sheet.cell(case_id+1, 4).value.find('${normal_tel}')!=-1:
                        print(row_data["check_sql"])
                        row_data["check_sql"]=DoRegs.do_regs(sheet.cell(case_id+1, 4).value)
                    else:
                        row_data["check_sql"] = sheet.cell(case_id+1, 4).value
                    row_data["expected"] = sheet.cell(case_id+1, 5).value  # 添加一个期望值到测试数据里面去
                    row_data["method"] = sheet.cell(case_id+1, 6).value
                    row_data["title"] = sheet.cell(case_id+1, 7).value
                    row_data["sheet_name"]=key
                    test_data.append(row_data)
                    cls.update_tel(normal_tel+2,file_name,"init") #更新手机号  针对excel的操作

        return test_data

    @staticmethod
    def write_back(file_name,sheet_name,row,col,result):#专门写会数据
        wb= load_workbook(file_name)
        sheet=wb[sheet_name]
        
        sheet.cell(row,col).value=result
        wb.save(file_name)#保存结果
        wb.close()

    @classmethod
    def update_tel(cls,tel,file_name,sheet_name): #更新excel里面手机号的数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(3,1).value=tel
        wb.save(file_name)




if __name__ == '__main__':
    res=Doexcel().get_excel(project_path.test_case_path)
    print(res)
